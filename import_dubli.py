"""
Import skipped duplicate clients from dubli.xlsx and dubli_list3.xlsx
- dubli.xlsx      → group "Himalay"        (rows marked "пропущен")
- dubli_list3.xlsx → group "Himalay ishxona" (all rows)
Phone field gets a unique placeholder so it can be filled in manually later.
"""
import openpyxl
import uuid
import psycopg2
from datetime import datetime, timezone

TENANT_ID    = '0f92d89b-e7e7-411c-8c89-b5dd801bbe6a'
GROUP_HIMALAY        = 'a6bfecf9-3d9f-4fbc-8bf8-0c3871fff8de'
GROUP_HIMALAY_ISHXONA = '8b2f3db8-59bc-486a-b852-58e489fd6399'

conn = psycopg2.connect(
    host='localhost', port=5432,
    dbname='suvpro', user='postgres', password='strongpassword'
)
cur = conn.cursor()

counter = 1  # global counter for unique placeholder phones

def make_phone(counter):
    return f'+0000{counter:05d}'   # e.g. +0000000001 — clearly fake

def insert_client(address, qty, group_id, row_num, note=''):
    global counter
    client_id = str(uuid.uuid4())
    phone = make_phone(counter)
    counter += 1

    cur.execute("""
        INSERT INTO clients
            (id, tenant_id, first_name, phone, is_active, is_verified, is_blocked,
             has_contract, container_balance, debt_amount, advance_amount,
             notes, group_id, created_at, updated_at, is_deleted)
        VALUES
            (%s, %s, %s, %s, true, false, false,
             false, %s, 0, 0,
             %s, %s, now(), now(), false)
    """, (
        client_id, TENANT_ID, address, phone,
        qty,
        f'Dubli import (qator {row_num}){(" — " + note) if note else ""}',
        group_id
    ))

    addr_id = str(uuid.uuid4())
    cur.execute("""
        INSERT INTO client_addresses
            (id, tenant_id, client_id, label, address_text, is_primary, created_at, updated_at)
        VALUES
            (%s, %s, %s, %s, %s, true, now(), now())
    """, (addr_id, TENANT_ID, client_id, 'Asosiy', address))

    return phone

# ── dubli.xlsx → Himalay ─────────────────────────────────────────────────────
wb1 = openpyxl.load_workbook('/root/suvpro/dubli.xlsx')
ws1 = wb1.active

added_himalay = []
for row in ws1.iter_rows(min_row=2, values_only=True):
    row_num, address, qty, dolg, phone_orig, note = row
    if not note or 'пропущен' not in str(note):
        continue
    ph = insert_client(address, qty or 0, GROUP_HIMALAY, row_num)
    added_himalay.append((row_num, address, qty, ph))

# ── dubli_list3.xlsx → Himalay ishxona ───────────────────────────────────────
wb2 = openpyxl.load_workbook('/root/suvpro/dubli_list3.xlsx')
ws2 = wb2.active

added_ishxona = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    row_num, address, qty, phone_orig, group_txt, addr_in_db = row
    if not address:
        continue
    ph = insert_client(address, qty or 0, GROUP_HIMALAY_ISHXONA, row_num)
    added_ishxona.append((row_num, address, qty, ph))

conn.commit()
cur.close()
conn.close()

print(f"\n✅ Himalay (dubli.xlsx) — {len(added_himalay)} ta mijoz qo'shildi:")
for r, a, q, p in added_himalay:
    print(f"  qator {r}: {a} | {q} ta | tel: {p}")

print(f"\n✅ Himalay ishxona (dubli_list3.xlsx) — {len(added_ishxona)} ta mijoz qo'shildi:")
for r, a, q, p in added_ishxona:
    print(f"  qator {r}: {a} | {q} ta | tel: {p}")

print(f"\nJami: {len(added_himalay) + len(added_ishxona)} ta mijoz")
print("⚠️  Telefon raqamlari '+0000XXXXX' ko'rinishida — keyinchalik qo'lda to'ldiring")
