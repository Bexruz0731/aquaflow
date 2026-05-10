from typing import Optional
from datetime import date, datetime, timezone, timedelta
from uuid import UUID
from fastapi import APIRouter, Depends, Query, Response, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, exists, and_, or_
from sqlalchemy.orm import selectinload

from app.db.base import get_db
from app.core.deps import require_operator
from app.models.user import User
from app.models.order import Order, OrderStatus, OrderItem
from app.models.client import Client, ClientAddress, ClientGroup
from app.models.courier import Courier
from app.models.finance import Debt
from app.models.product import Product

router = APIRouter()


@router.get("/orders")
async def orders_report(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50),
    user: User = Depends(require_operator),
    db: AsyncSession = Depends(get_db),
):
    query = select(Order).where(Order.tenant_id == user.tenant_id, Order.is_deleted == False)
    if date_from:
        query = query.where(Order.created_at >= datetime(date_from.year, date_from.month, date_from.day, tzinfo=timezone.utc))
    if date_to:
        query = query.where(Order.created_at <= datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59, tzinfo=timezone.utc))

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    # Summary — same filters as the paginated query
    summary_result = await db.execute(
        select(
            func.coalesce(func.sum(Order.total_amount - Order.discount_amount), 0),
            func.count(Order.id),
        ).select_from(query.subquery())
    )
    total_revenue, total_count = summary_result.one()

    result = await db.execute(
        query.order_by(Order.created_at.desc())
        .offset((page - 1) * per_page).limit(per_page)
        .options(
            selectinload(Order.client).selectinload(Client.addresses),
            selectinload(Order.courier).selectinload(Courier.user),
        )
    )
    orders = result.scalars().all()

    avg_check = (total_revenue or 0) // total_count if total_count else 0

    def _fmt(o: Order) -> dict:
        client_name = None
        if o.client:
            primary = next((a for a in o.client.addresses if a.is_primary), None) or (o.client.addresses[0] if o.client.addresses else None)
            if primary and primary.address_text:
                client_name = primary.address_text
            else:
                client_name = o.client.phone or None
        elif o.is_walkin:
            client_name = o.walkin_store or o.walkin_address or "Tez sotuv"
        courier_name = None
        if o.courier and o.courier.user:
            courier_name = f"{o.courier.user.first_name} {o.courier.user.last_name or ''}".strip()
        return {
            "id": o.id,
            "client_name": client_name,
            "courier_name": courier_name,
            "status": o.status,
            "total_amount": o.total_amount,
            "payment_status": o.payment_status,
            "created_at": o.created_at.isoformat() if o.created_at else None,
            "completed_at": o.delivered_at.isoformat() if o.delivered_at else None,
        }

    return {
        "summary": {
            "revenue": total_revenue or 0,
            "count": total_count or 0,
            "avg_check": avg_check,
        },
        "items": [_fmt(o) for o in orders],
        "total": total,
        "page": page,
    }


STATUS_LABELS = {
    "yangi": "Yangi",
    "tayinlandi": "Tayinlandi",
    "yetkazildi": "Yetkazildi",
    "bekor": "Bekor qilindi",
}
PAYMENT_STATUS_LABELS = {
    "tolanmagan": "To'lanmagan",
    "tolangan": "To'langan",
    "qisman": "Qisman to'langan",
}
PAYMENT_METHOD_LABELS = {
    "naqd": "Naqd",
    "karta": "Karta",
    "online": "Online",
    "": "—",
}


def _status_label(val: str) -> str:
    if val is None:
        return "—"
    key = str(val).lower().split(".")[-1]
    return STATUS_LABELS.get(key, val)


def _pay_status_label(val: str) -> str:
    if val is None:
        return "—"
    key = str(val).lower().split(".")[-1]
    return PAYMENT_STATUS_LABELS.get(key, val)


def _pay_method_label(val) -> str:
    if val is None:
        return "—"
    key = str(val).lower().split(".")[-1]
    return PAYMENT_METHOD_LABELS.get(key, str(val))


@router.get("/export/excel")
async def export_excel(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    sheets: str = Query("orders,clients,debts", description="Comma-separated sheet names"),
    user: User = Depends(require_operator),
    db: AsyncSession = Depends(get_db),
):
    """Generate Excel report with selected sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    requested = [s.strip() for s in sheets.split(",")]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 20

    # Sheet: Orders
    if "orders" in requested:
        ws = wb.create_sheet("Buyurtmalar")
        ws.append(["№", "Sana", "Mijoz FIO", "Mijoz telefon", "Kuryer", "Holat", "To'lov holati", "To'lov turi", "Jami (so'm)", "Yetkazildi"])
        style_header(ws)

        q = (
            select(Order)
            .where(Order.tenant_id == user.tenant_id, Order.is_deleted == False)
            .options(
                selectinload(Order.client),
                selectinload(Order.courier).selectinload(Courier.user),
            )
        )
        if date_from:
            q = q.where(Order.created_at >= datetime(date_from.year, date_from.month, date_from.day, tzinfo=timezone.utc))
        if date_to:
            q = q.where(Order.created_at <= datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59, tzinfo=timezone.utc))

        result = await db.execute(q.order_by(Order.created_at.desc()).limit(5000))
        for idx, o in enumerate(result.scalars().all(), 1):
            if o.client:
                client_name = f"{o.client.first_name} {o.client.last_name or ''}".strip()
                client_phone = o.client.phone
            elif o.is_walkin:
                client_name = o.walkin_store or o.walkin_address or "Tez sotuv"
                client_phone = o.walkin_phone or "—"
            else:
                client_name = "—"
                client_phone = "—"
            courier_name = (
                f"{o.courier.user.first_name} {o.courier.user.last_name or ''}".strip()
                if o.courier and o.courier.user else "—"
            )
            sana = o.created_at.strftime("%d.%m.%Y %H:%M") if o.created_at else "—"
            yetkazildi = o.delivered_at.strftime("%d.%m.%Y %H:%M") if o.delivered_at else "—"
            ws.append([
                idx, sana, client_name, client_phone, courier_name,
                _status_label(o.status),
                _pay_status_label(o.payment_status),
                _pay_method_label(o.payment_method),
                o.total_amount,
                yetkazildi,
            ])

        # Column widths
        for col, width in zip("ABCDEFGHIJ", [5, 18, 22, 16, 20, 14, 20, 14, 14, 18]):
            ws.column_dimensions[col].width = width

    # Sheet: Clients
    if "clients" in requested:
        ws = wb.create_sheet("Mijozlar")
        ws.append(["№", "Ism", "Familiya", "Telefon", "Qarz (so'm)", "Idish qarzi", "Holat"])
        style_header(ws)
        result = await db.execute(select(Client).where(Client.tenant_id == user.tenant_id).order_by(Client.created_at.desc()).limit(5000))
        for idx, c in enumerate(result.scalars().all(), 1):
            holat = "Bloklangan" if c.is_blocked else ("Faol" if c.is_active else "Nofaol")
            ws.append([idx, c.first_name, c.last_name or "", c.phone, c.debt_amount, c.container_balance, holat])
        for col, width in zip("ABCDEFG", [5, 18, 18, 16, 14, 12, 12]):
            ws.column_dimensions[col].width = width

    # Sheet: Debts
    if "debts" in requested:
        ws = wb.create_sheet("Qarzlar")
        ws.append(["№", "Mijoz FIO", "Telefon", "Asl qarz (so'm)", "Qolgan qarz (so'm)", "To'landi"])
        style_header(ws)
        result = await db.execute(
            select(Debt).where(Debt.tenant_id == user.tenant_id)
            .options(selectinload(Debt.client))
            .limit(5000)
        )
        for idx, d in enumerate(result.scalars().all(), 1):
            client_name = f"{d.client.first_name} {d.client.last_name or ''}".strip() if d.client else "—"
            phone = d.client.phone if d.client else "—"
            ws.append([idx, client_name, phone, d.original_amount, d.remaining_amount, "Ha" if d.is_paid else "Yo'q"])
        for col, width in zip("ABCDEF", [5, 22, 16, 18, 18, 10]):
            ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"suvpro_hisobot_{date_from or 'boshidan'}_{date_to or 'hozir'}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/products")
async def products_report(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    user: User = Depends(require_operator),
    db: AsyncSession = Depends(get_db),
):
    """Get sales report by product."""
    # Build base query for order items with date filters
    item_query = (
        select(OrderItem)
        .join(Order, OrderItem.order_id == Order.id)
        .where(
            OrderItem.tenant_id == user.tenant_id,
            Order.is_deleted == False,
            Order.status.in_([OrderStatus.YETKAZILDI, OrderStatus.YOPILDI])
        )
    )

    if date_from:
        item_query = item_query.where(
            Order.created_at >= datetime(date_from.year, date_from.month, date_from.day, tzinfo=timezone.utc)
        )
    if date_to:
        item_query = item_query.where(
            Order.created_at <= datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59, tzinfo=timezone.utc)
        )

    # Get all products for this tenant
    products_result = await db.execute(
        select(Product).where(Product.tenant_id == user.tenant_id)
    )
    products = {p.id: p for p in products_result.scalars().all()}

    # Get sales data grouped by product
    from sqlalchemy import case
    sales_query = (
        select(
            OrderItem.product_id,
            func.sum(OrderItem.quantity).label('total_quantity'),
            func.sum(OrderItem.total).label('total_revenue'),
            func.count(func.distinct(OrderItem.order_id)).label('order_count')
        )
        .select_from(OrderItem)
        .join(Order, OrderItem.order_id == Order.id)
        .where(
            OrderItem.tenant_id == user.tenant_id,
            Order.is_deleted == False,
            Order.status.in_([OrderStatus.YETKAZILDI, OrderStatus.YOPILDI])
        )
        .group_by(OrderItem.product_id)
    )

    if date_from:
        sales_query = sales_query.where(
            Order.created_at >= datetime(date_from.year, date_from.month, date_from.day, tzinfo=timezone.utc)
        )
    if date_to:
        sales_query = sales_query.where(
            Order.created_at <= datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59, tzinfo=timezone.utc)
        )

    sales_result = await db.execute(sales_query)
    sales_data = sales_result.all()

    # Build response
    product_stats = []
    for product_id, quantity, revenue, order_count in sales_data:
        product = products.get(product_id)
        if product:
            product_stats.append({
                "product_id": str(product_id),
                "product_name": product.name,
                "total_quantity": quantity or 0,
                "total_revenue": revenue or 0,
                "order_count": order_count or 0,
                "avg_price": (revenue // quantity) if quantity else 0,
            })

    # Sort by revenue descending
    product_stats.sort(key=lambda x: x["total_revenue"], reverse=True)

    # Calculate totals
    total_quantity = sum(p["total_quantity"] for p in product_stats)
    total_revenue = sum(p["total_revenue"] for p in product_stats)

    return {
        "summary": {
            "total_quantity": total_quantity,
            "total_revenue": total_revenue,
            "product_count": len(product_stats),
        },
        "products": product_stats,
    }


@router.get("/export/product/{product_id}")
async def export_product_detail(
    product_id: str,
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    user: User = Depends(require_operator),
    db: AsyncSession = Depends(get_db),
):
    """Export detailed order information for a specific product to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from uuid import UUID

    try:
        product_uuid = UUID(product_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid product ID")

    # Get product info
    product_result = await db.execute(
        select(Product).where(Product.id == product_uuid, Product.tenant_id == user.tenant_id)
    )
    product = product_result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Build query for orders containing this product
    query = (
        select(Order, OrderItem, Client, Courier)
        .join(OrderItem, OrderItem.order_id == Order.id)
        .outerjoin(Client, Client.id == Order.client_id)
        .outerjoin(Courier, Courier.id == Order.courier_id)
        .where(
            OrderItem.product_id == product_uuid,
            OrderItem.tenant_id == user.tenant_id,
            Order.is_deleted == False,
            Order.status.in_([OrderStatus.YETKAZILDI, OrderStatus.YOPILDI])
        )
        .options(
            selectinload(Courier.user)
        )
    )

    if date_from:
        query = query.where(
            Order.created_at >= datetime(date_from.year, date_from.month, date_from.day, tzinfo=timezone.utc)
        )
    if date_to:
        query = query.where(
            Order.created_at <= datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59, tzinfo=timezone.utc)
        )

    result = await db.execute(query.order_by(Order.created_at.desc()).limit(5000))
    orders_data = result.all()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = product.name[:31]  # Excel sheet name limit

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "№", "Buyurtma №", "Sana", "Mijoz", "Mijoz telefon",
        "Kuryer", "Miqdor", "Narx", "Jami", "To'lov holati", "Yetkazildi"
    ]
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 20

    # Data rows
    total_quantity = 0
    total_revenue = 0

    for idx, (order, order_item, client, courier) in enumerate(orders_data, 1):
        client_name = f"{client.first_name} {client.last_name or ''}".strip() if client else "—"
        client_phone = client.phone if client else "—"
        courier_name = "—"
        if courier and courier.user:
            courier_name = f"{courier.user.first_name} {courier.user.last_name or ''}".strip()

        created_at = order.created_at.strftime("%d.%m.%Y %H:%M") if order.created_at else "—"
        delivered_at = order.delivered_at.strftime("%d.%m.%Y %H:%M") if order.delivered_at else "—"

        total_quantity += order_item.quantity
        total_revenue += order_item.total

        ws.append([
            idx,
            order.id,
            created_at,
            client_name,
            client_phone,
            courier_name,
            order_item.quantity,
            order_item.price_at_order,
            order_item.total,
            _pay_status_label(order.payment_status),
            delivered_at,
        ])

    # Summary row
    ws.append([])
    summary_row = ws.max_row + 1
    ws.append([
        "", "", "", "", "", "JAMI:",
        total_quantity, "", total_revenue, "", ""
    ])

    # Style summary row
    for cell in ws[summary_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")

    # Column widths
    for col, width in zip("ABCDEFGHIJK", [5, 12, 16, 22, 16, 20, 10, 12, 14, 16, 16]):
        ws.column_dimensions[col].width = width

    # Generate file
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{product.name}_hisobot_{date_from or 'boshidan'}_{date_to or 'hozir'}.xlsx"
    # Clean filename from invalid characters
    filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()

    # Encode filename for Content-Disposition header to support non-ASCII characters
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )


@router.get("/inactive-by-product")
async def inactive_by_product(
    product_id: UUID = Query(...),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    user: User = Depends(require_operator),
    db: AsyncSession = Depends(get_db),
):
    """
    Clients who have ordered a specific product at least once
    but NOT in the last product.inactive_threshold_days days.
    """
    # Load product and its threshold
    prod_result = await db.execute(
        select(Product).where(Product.id == product_id, Product.tenant_id == user.tenant_id)
    )
    product = prod_result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    threshold_days = product.inactive_threshold_days
    cutoff = datetime.now(timezone.utc) - timedelta(days=threshold_days)

    # Only count delivered orders (not cancelled/pending)
    delivered_filter = and_(
        OrderItem.product_id == product_id,
        Order.tenant_id == user.tenant_id,
        Order.is_deleted == False,
        Order.status == OrderStatus.YETKAZILDI,
        Order.client_id.isnot(None),
    )

    # Subquery: clients who have EVER received this product
    ever_ordered_sq = (
        select(Order.client_id)
        .join(OrderItem, OrderItem.order_id == Order.id)
        .where(delivered_filter)
        .distinct()
        .subquery()
    )

    # Subquery: last delivered order date for this product per client
    last_product_order_sq = (
        select(
            Order.client_id,
            func.max(Order.created_at).label("last_at"),
        )
        .join(OrderItem, OrderItem.order_id == Order.id)
        .where(delivered_filter)
        .group_by(Order.client_id)
        .subquery()
    )

    # Subquery: delivered order count for this product per client
    product_order_count_sq = (
        select(
            Order.client_id,
            func.count(func.distinct(Order.id)).label("product_orders"),
        )
        .join(OrderItem, OrderItem.order_id == Order.id)
        .where(delivered_filter)
        .group_by(Order.client_id)
        .subquery()
    )

    # Base query: active clients who ever received this product but NOT in the last N days
    query = (
        select(Client)
        .join(ever_ordered_sq, ever_ordered_sq.c.client_id == Client.id)
        .outerjoin(last_product_order_sq, last_product_order_sq.c.client_id == Client.id)
        .where(
            Client.tenant_id == user.tenant_id,
            Client.is_deleted == False,
            Client.is_active == True,
            last_product_order_sq.c.last_at < cutoff,
        )
    )

    if search:
        query = query.where(
            func.concat(Client.first_name, ' ', func.coalesce(Client.last_name, '')).ilike(f"%{search}%")
            | Client.phone.ilike(f"%{search}%")
        )

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    result = await db.execute(
        query
        .add_columns(
            last_product_order_sq.c.last_at,
            product_order_count_sq.c.product_orders,
        )
        .outerjoin(product_order_count_sq, product_order_count_sq.c.client_id == Client.id)
        .order_by(last_product_order_sq.c.last_at.asc())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    rows = result.all()

    # Bulk fetch primary addresses
    client_ids = [r[0].id for r in rows]
    addr_map: dict = {}
    if client_ids:
        addr_result = await db.execute(
            select(ClientAddress)
            .where(ClientAddress.client_id.in_(client_ids), ClientAddress.is_primary == True)
        )
        for addr in addr_result.scalars().all():
            addr_map[addr.client_id] = addr

    items = []
    for client, last_at, product_orders in rows:
        primary_addr = addr_map.get(client.id)
        items.append({
            "id": str(client.id),
            "first_name": client.first_name,
            "last_name": client.last_name,
            "phone": client.phone,
            "container_balance": client.container_balance,
            "debt_amount": client.debt_amount,
            "last_product_order_at": last_at.isoformat() if last_at else None,
            "product_orders_count": product_orders or 0,
            "address": primary_addr.address_text if primary_addr else None,
            "address_label": primary_addr.label if primary_addr else None,
        })

    return {
        "product": {
            "id": str(product.id),
            "name": product.name,
            "inactive_threshold_days": threshold_days,
            "is_returnable_container": product.is_returnable_container,
        },
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.get("/inactive-by-group")
async def inactive_by_group(
    group_id: Optional[UUID] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    user: User = Depends(require_operator),
    db: AsyncSession = Depends(get_db),
):
    """
    Clients in a specific group (or ungrouped if group_id is None)
    who have NOT placed any delivered order in the last group.inactive_threshold_days days.
    """
    threshold_days = 30
    group_name = "Guruхsiz"

    if group_id is not None:
        grp_result = await db.execute(
            select(ClientGroup).where(ClientGroup.id == group_id, ClientGroup.tenant_id == user.tenant_id)
        )
        group = grp_result.scalar_one_or_none()
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        threshold_days = group.inactive_threshold_days
        group_name = group.name

    cutoff = datetime.now(timezone.utc) - timedelta(days=threshold_days)

    last_order_sq = (
        select(
            Order.client_id,
            func.max(Order.created_at).label("last_at"),
        )
        .where(
            Order.tenant_id == user.tenant_id,
            Order.is_deleted == False,
            Order.status == OrderStatus.YETKAZILDI,
            Order.client_id.isnot(None),
        )
        .group_by(Order.client_id)
        .subquery()
    )

    order_count_sq = (
        select(
            Order.client_id,
            func.count(func.distinct(Order.id)).label("total_orders"),
        )
        .where(
            Order.tenant_id == user.tenant_id,
            Order.is_deleted == False,
            Order.status == OrderStatus.YETKAZILDI,
            Order.client_id.isnot(None),
        )
        .group_by(Order.client_id)
        .subquery()
    )

    if group_id is not None:
        group_filter = Client.group_id == group_id
    else:
        group_filter = Client.group_id.is_(None)

    query = (
        select(Client)
        .join(last_order_sq, last_order_sq.c.client_id == Client.id)
        .where(
            Client.tenant_id == user.tenant_id,
            Client.is_deleted == False,
            Client.is_active == True,
            group_filter,
            last_order_sq.c.last_at < cutoff,
        )
    )

    if search:
        safe = search.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        query = query.where(
            func.concat(Client.first_name, ' ', func.coalesce(Client.last_name, '')).ilike(f"%{safe}%")
            | Client.phone.ilike(f"%{safe}%")
        )

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    result = await db.execute(
        query
        .add_columns(last_order_sq.c.last_at, order_count_sq.c.total_orders)
        .outerjoin(order_count_sq, order_count_sq.c.client_id == Client.id)
        .order_by(last_order_sq.c.last_at.asc().nullsfirst())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    rows = result.all()

    client_ids = [r[0].id for r in rows]
    addr_map: dict = {}
    if client_ids:
        addr_result = await db.execute(
            select(ClientAddress)
            .where(ClientAddress.client_id.in_(client_ids), ClientAddress.is_primary == True)
        )
        for addr in addr_result.scalars().all():
            addr_map[addr.client_id] = addr

    items = []
    for client, last_at, total_orders in rows:
        primary_addr = addr_map.get(client.id)
        items.append({
            "id": str(client.id),
            "first_name": client.first_name,
            "last_name": client.last_name,
            "phone": client.phone,
            "container_balance": client.container_balance,
            "debt_amount": client.debt_amount,
            "last_order_at": last_at.isoformat() if last_at else None,
            "orders_count": total_orders or 0,
            "address": primary_addr.address_text if primary_addr else None,
            "address_label": primary_addr.label if primary_addr else None,
        })

    return {
        "group": {
            "id": str(group_id) if group_id else None,
            "name": group_name,
            "inactive_threshold_days": threshold_days,
        },
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.get("/never-ordered")
async def never_ordered_clients(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    user: User = Depends(require_operator),
    db: AsyncSession = Depends(get_db),
):
    """Clients who have never placed a single order."""
    from sqlalchemy import exists as sa_exists
    from app.models.order import Order

    has_order_sq = (
        select(Order.client_id)
        .where(Order.tenant_id == user.tenant_id, Order.is_deleted == False, Order.client_id == Client.id)
        .correlate(Client)
    )

    query = (
        select(Client)
        .where(
            Client.tenant_id == user.tenant_id,
            Client.is_deleted == False,
            Client.is_active == True,
            ~sa_exists(has_order_sq),
        )
    )

    if search:
        safe = search.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        addr_sq = (
            select(ClientAddress.client_id)
            .where(ClientAddress.address_text.ilike(f"%{safe}%"), ClientAddress.client_id == Client.id)
            .correlate(Client)
        )
        query = query.where(
            or_(
                Client.phone.ilike(f"%{safe}%"),
                Client.first_name.ilike(f"%{safe}%"),
                sa_exists(addr_sq),
            )
        )

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    result = await db.execute(
        query.order_by(Client.created_at.desc()).offset((page - 1) * per_page).limit(per_page)
    )
    clients = result.scalars().all()

    client_ids = [c.id for c in clients]
    addr_map: dict = {}
    if client_ids:
        addr_result = await db.execute(
            select(ClientAddress).where(ClientAddress.client_id.in_(client_ids), ClientAddress.is_primary == True)
        )
        for addr in addr_result.scalars().all():
            addr_map[addr.client_id] = addr

    items = []
    for c in clients:
        primary_addr = addr_map.get(c.id)
        items.append({
            "id": str(c.id),
            "first_name": c.first_name,
            "last_name": c.last_name,
            "phone": c.phone,
            "debt_amount": c.debt_amount,
            "address": primary_addr.address_text if primary_addr else None,
            "registered_at": c.created_at.isoformat() if c.created_at else None,
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }
